# -*- coding: utf-8 -*-

# Abandon 写clear每个输入框的按钮（已经放了，但是没功能）
# Done: 冰冰姐的建议：1. A+B->P初始浓度不同时，最好在前台明确提醒一下算的是哪个物质的反应时间。
# Done: 冰冰姐的建议：2. 强烈建议增加kcal/mol单位
# Abandon: 冰冰姐的建议：3. 能垒与速率常数可以考虑旁边设一个toggle,因为不可能同时设置这两个参数的
# Done: 让程序记住上次算的是哪两个参数，比如k和conv。
#  新增一个按钮，按钮的文本随上次计算的参数变化而变化，比如上次计算的是k和conv，按钮的文本就是Recalculate k & conv，按钮在计算前会清空这两个框
#  然后要做的就是改一个浓度，点这个按钮；改一个浓度，点这个按钮
# Done: 支持并测试batch

__author__ = 'LiYuanhe'

from datetime import datetime
import openpyxl
from Python_Lib.My_Lib_PyQt6 import *
from Python_Lib.My_Lib import read_xlsx, write_xlsx
from Lib import *

if not QtWidgets.QApplication.instance():
    Application = QtWidgets.QApplication(sys.argv)
    if platform.system() == 'Windows':
        import ctypes

        APPID = 'LYH.EyringEq.0.1'
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(APPID)
        Application.setWindowIcon(QtGui.QIcon('UI/Eyring_Eq.png'))

pyqt_ui_compile('Eyring_Eq.py')
from UI.Eyring_Eq import Ui_Eyring_Eq


def evaluate_expression(expression: str) -> Optional[float]:
    if "~" in expression:  # 处理近似为100表示为“~100”时，程序认为“~”是bitwise not操作符的问题。
        return None
    expression = expression.replace(' × 10^', 'E')
    try:
        ret = eval(expression)
        if is_float(ret):
            return ret
    except Exception:
        return None


class myWidget(Ui_Eyring_Eq, QtWidgets.QWidget, Qt_Widget_Common_Functions):

    def __init__(self):
        super(self.__class__, self).__init__()
        self.setupUi(self)
        self.setMinimumWidth(550)
        self.setMinimumHeight(366)

        self.k1_UNIT_HTML = "<html><head/><body><p>s<span style=\" vertical-align:super;\">-1</span></p></body></html>"
        self.k2_UNIT_HTML = "<html><head/><body><p>M<span style=\" vertical-align:super;\">-1</span>·s<span style=\" vertical-align:super;\">-1</span></p></body></html>"

        self.MODE_MAPPING = {"AB": self.bimolecular_AB_radioButton,
                             "AA": self.bimolecular_AA_radioButton,
                             "Acat": self.bimolecular_A_Cat_radioButton,
                             "A": self.unimolecular_radioButton}
        self.INPUT_MAPPING: Dict[str, Qt.QLineEdit] = OrderedDict([("G", self.G_neq_lineEdit),
                                                                   ("T", self.temp_lineEdit),
                                                                   ("σ", self.sigma_lineEdit),
                                                                   ("kTST", self.kTST_lineEdit),
                                                                   ("conc1", self.conc1_lineEdit),
                                                                   ("conc2", self.conc2_lineEdit),
                                                                   ("conv", self.conversion_lineEdit),
                                                                   ("t", self.total_time_lineEdit)])
        self.RECALC_MAPPING = {"G":"ΔG≠",
                               "T":"temp.",
                               "t":"time",
                               "conv":"conv.",
                               "kTST":"kTST"}

        connect_once(self.unimolecular_radioButton, self.mode_change)
        connect_once(self.bimolecular_AA_radioButton, self.mode_change)
        connect_once(self.bimolecular_AB_radioButton, self.mode_change)
        connect_once(self.bimolecular_A_Cat_radioButton, self.mode_change)
        connect_once(self.reset_all_pushButton, self.reset_all)
        connect_once(self.calculate_pushButton, self.calc)
        connect_once(self.batch_pushButton,self.batch)
        connect_once(self.recalc_pushButton,self.recalc)

        connect_once(self.conversion_lineEdit, self.check_fill_status)
        connect_once(self.total_time_lineEdit, self.check_fill_status)
        connect_once(self.temp_lineEdit, self.check_fill_status)
        connect_once(self.G_neq_lineEdit, self.check_fill_status)
        connect_once(self.kTST_lineEdit, self.check_fill_status)
        connect_once(self.conc2_lineEdit, self.check_fill_status)
        connect_once(self.conc1_lineEdit, self.check_fill_status)
        connect_once(self.sigma_lineEdit, self.check_fill_status)
        connect_once(self.total_time_lineEdit, self.smart_display_total_time)

        connect_once(self.temp_lineEdit, self.clear_kTST)
        connect_once(self.G_neq_lineEdit, self.clear_kTST)
        connect_once(self.sigma_lineEdit, self.clear_kTST)

        connect_once(self.energy_unit_comboBox.currentTextChanged, self.G_neq_unit_changed)
        connect_once(self.time_unit_comboBox.currentTextChanged, self.time_unit_changed)

        self.clear_conc1_pushButton.hide()
        self.clear_conc2_pushButton.hide()
        self.clear_G_pushButton.hide()
        self.clear_k_pushButton.hide()
        self.clear_T_pushButton.hide()
        self.clear_t_pushButton.hide()
        self.clear_conv_pushButton.hide()
        self.clear_sigma_pushButton.hide()

        self.energy_unit_comboBox_before_change = self.energy_unit_comboBox.currentText()
        self.time_unit_comboBox_before_change = self.time_unit_comboBox.currentText()

        self.last_unknowns = [] # remember which parameters to be recalculated

        # This is not a space.
        # This is an unicode blank to tell the program that the kTST is calculated, instead of user input.
        self.kTST_is_calculated_marker = "⠀"

        # 动一下，触发信号
        self.unimolecular_radioButton.click()
        self.check_fill_status()
        self.show()
        self.center_the_widget()
        self.resize(self.minimumSizeHint())

    # def clear_G(self):
    #     self.G_neq_lineEdit.setText("")

    def mode_change(self):
        self.check_fill_status()
        self.clear_kTST()
        if self.sender() == self.unimolecular_radioButton:
            print("Unimolecular Mode")
            self.conc1_lineEdit.hide()
            self.conc2_lineEdit.hide()
            self.conc1_label.hide()
            self.conc2_label.hide()
            self.conc1_unit_label.hide()
            self.conc2_unit_label.hide()
            self.conc1_lineEdit.setText("")
            self.conc2_lineEdit.setText("")
            self.conc1_lineEdit.setEnabled(False)
            self.conc2_lineEdit.setEnabled(False)
            self.conc1_lineEdit.setFocusPolicy(QtCore.Qt.FocusPolicy.NoFocus)
            self.conc2_lineEdit.setFocusPolicy(QtCore.Qt.FocusPolicy.NoFocus)
            self.kTST_unit_label.setText(self.k1_UNIT_HTML)

        elif self.sender() == self.bimolecular_AA_radioButton:
            print("A+A mode")
            self.conc1_lineEdit.show()
            self.conc2_lineEdit.hide()
            self.conc1_label.show()
            self.conc2_label.hide()
            self.conc1_unit_label.show()
            self.conc2_unit_label.hide()
            self.conc1_lineEdit.setEnabled(True)
            self.conc2_lineEdit.setEnabled(False)
            self.conc1_lineEdit.setFocusPolicy(QtCore.Qt.FocusPolicy.StrongFocus)
            self.conc2_lineEdit.setFocusPolicy(QtCore.Qt.FocusPolicy.NoFocus)
            self.conc2_lineEdit.setText("")
            self.conc2_label.setText("Conc. A")
            self.kTST_unit_label.setText(self.k2_UNIT_HTML)

        elif self.sender() == self.bimolecular_AB_radioButton:
            print("A+B mode")
            self.conc1_lineEdit.show()
            self.conc2_lineEdit.show()
            self.conc1_label.show()
            self.conc2_label.show()
            self.conc1_unit_label.show()
            self.conc2_unit_label.show()
            self.conc1_lineEdit.setEnabled(True)
            self.conc2_lineEdit.setEnabled(True)
            self.conc1_lineEdit.setFocusPolicy(QtCore.Qt.FocusPolicy.StrongFocus)
            self.conc2_lineEdit.setFocusPolicy(QtCore.Qt.FocusPolicy.StrongFocus)
            self.kTST_unit_label.setText(self.k2_UNIT_HTML)
            self.conc2_label.setText("Conc. B")

        elif self.sender() == self.bimolecular_A_Cat_radioButton:
            print("A+Cat mode")
            self.conc1_lineEdit.hide()
            self.conc2_lineEdit.show()
            self.conc1_label.hide()
            self.conc2_label.show()
            self.conc1_unit_label.hide()
            self.conc2_unit_label.show()
            self.conc1_lineEdit.setEnabled(False)
            self.conc2_lineEdit.setEnabled(True)
            self.conc1_lineEdit.setFocusPolicy(QtCore.Qt.FocusPolicy.NoFocus)
            self.conc2_lineEdit.setFocusPolicy(QtCore.Qt.FocusPolicy.StrongFocus)
            self.kTST_unit_label.setText(self.k2_UNIT_HTML)
            self.conc2_label.setText("Conc. Cat")

    def reset_all(self):
        self.G_neq_lineEdit.setText("")
        self.energy_unit_comboBox.setCurrentText("kJ/mol")
        self.temp_lineEdit.setText("")
        self.conc1_lineEdit.setText("")
        self.conc2_lineEdit.setText("")
        self.sigma_lineEdit.setText("1")
        self.total_time_lineEdit.setText("")
        self.time_unit_comboBox.setCurrentText("s")
        self.conversion_lineEdit.setText("98")
        # self.unimolecular_radioButton.click()

    def clear_kTST(self):
        if self.kTST_lineEdit.text().startswith(self.kTST_is_calculated_marker):
            self.kTST_lineEdit.setText("")

    def smart_display_total_time(self):
        if self.data["t"] and self.data["t"] > 60:
            self.total_time_display_label.setText("= " + smart_print_time(self.data["t"]))
        else:
            self.total_time_display_label.setText("")

    def G_neq_unit_changed(self, energy_unit):
        if self.data['G'] is None:
            ret_lineEdit_content = ""
        elif energy_unit == 'kJ/mol':
            ret_lineEdit_content = self.data['G'] / 1000
        elif energy_unit == "kcal/mol":
            ret_lineEdit_content = self.data['G'] / 1000 / kcal__kJ
        elif energy_unit == "eV":
            ret_lineEdit_content = self.data['G'] / 1000 / eV__kJ
        else:
            raise Exception("ComboBox Error")

        # 统一保留5位有效数字
        if ret_lineEdit_content:
            if ret_lineEdit_content > 100000:
                ret_lineEdit_content = str(int(ret_lineEdit_content))
            else:
                ret_lineEdit_content = "{:.5g}".format(ret_lineEdit_content)

        self.G_neq_lineEdit.setText(ret_lineEdit_content)
        self.energy_unit_comboBox_before_change = energy_unit

    def time_unit_changed(self, t_unit):
        if self.data['t'] is None:
            ret_lineEdit_content = ""
        elif t_unit == 's':
            ret_lineEdit_content = self.data['t']
        elif t_unit == 'min':
            ret_lineEdit_content = self.data['t'] / 60
        elif t_unit == 'h':
            ret_lineEdit_content = self.data['t'] / 3600
        elif t_unit == 'd':
            ret_lineEdit_content = self.data['t'] / 86400
        elif t_unit == 'year':
            ret_lineEdit_content = self.data['t'] / 86400 / 365
        else:
            raise Exception("ComboBox Error")

        # 统一保留5位有效数字
        if ret_lineEdit_content:
            if ret_lineEdit_content > 100000:
                ret_lineEdit_content = str(int(ret_lineEdit_content))
            else:
                ret_lineEdit_content = "{:.4g}".format(ret_lineEdit_content)

        self.total_time_lineEdit.setText(ret_lineEdit_content)
        self.time_unit_comboBox_before_change = t_unit

    def set_kTST_lineEdit(self, kTST):
        if self.unimolecular_radioButton.isChecked():
            self.kTST_lineEdit.setText(self.kTST_is_calculated_marker + smart_format_float(kTST))
        else:
            self.kTST_lineEdit.setText(self.kTST_is_calculated_marker + smart_format_float(kTST * 1000))

    def check_reasonable_input(self):

        self.G_neq_lineEdit.setStyleSheet("")
        self.temp_lineEdit.setStyleSheet("")
        self.total_time_lineEdit.setStyleSheet("")
        self.sigma_lineEdit.setStyleSheet("")
        self.kTST_lineEdit.setStyleSheet("")
        self.conc1_lineEdit.setStyleSheet("")
        self.conc2_lineEdit.setStyleSheet("")
        self.conversion_lineEdit.setStyleSheet("")

        if self.data['G'] is not None and self.data['G'] <= 0:
            self.G_neq_lineEdit.setStyleSheet("background-color: rgb(255, 219, 219);")
            self.data['G'] = None

        if self.data['t'] is not None and self.data['t'] <= 0:
            self.total_time_lineEdit.setStyleSheet("background-color: rgb(255, 219, 219);")
            self.data['t'] = None

        if self.data['T'] is not None and self.data['T'] <= 1:
            self.temp_lineEdit.setStyleSheet("background-color: rgb(255, 219, 219);")
            self.data['T'] = None

        if self.data['c1'] is not None and self.data['c1'] <= 0:
            self.conc1_lineEdit.setStyleSheet("background-color: rgb(255, 219, 219);")
            self.data['c1'] = None

        if self.data['c2'] is not None and self.data['c2'] <= 0:
            self.conc2_lineEdit.setStyleSheet("background-color: rgb(255, 219, 219);")
            self.data['c2'] = None

        if self.data['conv'] is not None and not 0 < self.data['conv'] < 100:
            self.conversion_lineEdit.setStyleSheet("background-color: rgb(255, 219, 219);")
            self.data['conv'] = None

    def check_fill_status(self):
        self.data = {"G": evaluate_expression(self.G_neq_lineEdit.text()),
                     "T": evaluate_expression(self.temp_lineEdit.text()),
                     "c1": evaluate_expression(self.conc1_lineEdit.text()),
                     "c2": evaluate_expression(self.conc2_lineEdit.text()),
                     "t": evaluate_expression(self.total_time_lineEdit.text()),
                     "conv": evaluate_expression(self.conversion_lineEdit.text()),
                     "σ": evaluate_expression(self.sigma_lineEdit.text())}

        self.conv_label.setText("Conv. A")
        if self.bimolecular_AB_radioButton.isChecked():
            if self.data["c1"] is not None and self.data["c2"] is not None:
                if self.data["c1"] <= self.data["c2"]:
                    self.conv_label.setText("Conv. A")
                else:
                    self.conv_label.setText("Conv. B")

        # Change to SI units
        if self.data['G'] is not None:
            G_unit = self.energy_unit_comboBox.currentText()
            if G_unit == 'kJ/mol':
                self.data['G'] *= 1000  # kJ-> J
            elif G_unit == "kcal/mol":
                self.data['G'] *= 1000 * kcal__kJ  # kcal-> J
            elif G_unit == "eV":
                self.data['G'] *= 1000 * eV__kJ  # kcal-> J
            else:
                raise Exception("ComboBox Error")

        if self.data['t'] is not None:
            t_unit = self.time_unit_comboBox.currentText()
            if t_unit == 's':
                pass
            elif t_unit == 'min':
                self.data['t'] *= 60
            elif t_unit == 'h':
                self.data['t'] *= 3600
            elif t_unit == 'd':
                self.data['t'] *= 86400
            elif t_unit == 'year':
                self.data['t'] *= 86400 * 365
            else:
                raise Exception("ComboBox Error")

        if self.data['T'] is not None:
            self.data['T'] += 273

        if self.data['c1'] is not None:
            self.data['c1'] *= 1000  # mol/L -> mol/m^3

        if self.data['c2'] is not None:
            self.data['c2'] *= 1000  # mol/L -> mol/m^3

        if self.data['conv'] is not None:
            self.data['conv'] /= 100  # percent to number

        if self.kTST_lineEdit.text().startswith(self.kTST_is_calculated_marker):
            self.data['kTST'] = None
        else:
            self.data['kTST'] = evaluate_expression(self.kTST_lineEdit.text())
            if not self.unimolecular_radioButton.isChecked() and self.data['kTST']:
                self.data['kTST'] /= 1000  # L/mol·s --> m3/mol·s

        self.check_reasonable_input()

        self.is_None = set([key for key, value in self.data.items() if value is None])
        if self.unimolecular_radioButton.isChecked():
            self.is_None.discard("c2")
            self.is_None.discard("c1")
        elif self.bimolecular_AA_radioButton.isChecked():
            self.is_None.discard("c2")
        elif self.bimolecular_A_Cat_radioButton.isChecked():
            self.is_None.discard("c1")

        allowed_missing_situations = [["G", 'kTST'],
                                      ["T", 'kTST'],
                                      ["t", 'kTST'],
                                      ["conv", 'kTST'],
                                      ["G", 't'],
                                      ["G", 'conv'],
                                      ["T", 't'],
                                      ["T", 'conv']]

        calc_allowed = any([set(x) == self.is_None for x in allowed_missing_situations])
        self.calculate_pushButton.setEnabled(calc_allowed)

        if len(self.last_unknowns)==2:
            recalc_allowed = any([set(x) == set(list(self.is_None)+self.last_unknowns) for x in allowed_missing_situations])
            self.recalc_pushButton.setEnabled(recalc_allowed)
        else:
            self.recalc_pushButton.setEnabled(False)


    def calc(self):
        self.last_unknowns = []

        # 确定各模式的Eyring方程的delta-n, 指定各模式动力学的求解函数
        if self.unimolecular_radioButton.isChecked():
            Δn = 0

            def k_from_kinetics(conv, time, conc1=None, conc2=None):
                return first_order_k_TST(conv, time)

            def t_from_kinetics(kTST, conv, conc1=None, conc2=None):
                return first_order_reaction_time(kTST, conv)

            def conv_from_kinetics(kTST, time, conc1=None, conc2=None):
                return first_order_conversion(kTST, time)

        elif self.bimolecular_AA_radioButton.isChecked():
            Δn = 1

            def k_from_kinetics(conv, time, conc1, conc2=None):
                return second_order_k_TST_A_plus_A(conv, time, conc1)

            def t_from_kinetics(kTST, conv, conc1, conc2=None):
                return second_order_reaction_time_A_plus_A(kTST, conv, conc1)

            def conv_from_kinetics(kTST, time, conc1, conc2=None):
                return second_order_conv_A_plus_A(kTST, time, conc1)

        elif self.bimolecular_AB_radioButton.isChecked():
            Δn = 1

            def k_from_kinetics(conv, time, conc1, conc2):
                if conc1 == conc2:
                    return second_order_k_TST_A_plus_A(conv, time, conc1)
                else:
                    return second_order_k_TST_A_plus_B(conv, time, conc1, conc2)

            def t_from_kinetics(kTST, conv, conc1, conc2):
                if conc1 == conc2:
                    return second_order_reaction_time_A_plus_A(kTST, conv, conc1)
                else:
                    return second_order_reaction_time_A_plus_B(kTST, conv, conc1, conc2)

            def conv_from_kinetics(kTST, time, conc1, conc2):
                if conc1 == conc2:
                    return second_order_conv_A_plus_A(kTST, time, conc1)
                else:
                    return second_order_conv_A_plus_B(kTST, time, conc1, conc2)

        elif self.bimolecular_A_Cat_radioButton.isChecked():
            Δn = 1

            # 把浓度并入速率常数，变成一级动力学
            def k_from_kinetics(conv, time, conc1, conc_cat):
                return first_order_k_TST(conv, time) / conc_cat

            def t_from_kinetics(kTST, conv, conc1, conc_cat):
                return first_order_reaction_time(kTST * conc_cat, conv)

            def conv_from_kinetics(kTST, time, conc1, conc_cat):
                return first_order_conversion(kTST * conc_cat, time)
        else:
            raise Exception("None of the radioButtons is checked.")

        G, T, c1, c2, t, conv, σ, kTST = [self.data[key] for key in ["G", "T", "c1", "c2", "t", "conv", "σ", 'kTST']]

        # 知道动力学，从动力学逆推kTST
        if "t" not in self.is_None and 'conv' not in self.is_None:
            kTST = k_from_kinetics(conv, t, c1, c2)
            self.set_kTST_lineEdit(kTST)
            self.last_unknowns.append("kTST")

        if kTST:
            if "G" in self.is_None:  # 知道G不知道T
                G = solve_for_ΔG(kTST, Δn, σ, T)
                # 把能量时间单位换回kJ/mol，填入答案，再把单位切换到想要的单位，这样对应的slot会自动换算单位
                current_energy_unit = self.energy_unit_comboBox.currentText()
                self.energy_unit_comboBox.setCurrentText("kJ/mol")
                self.G_neq_lineEdit.setText(smart_format_float(G / 1000, precision=4))
                self.last_unknowns.append("G")
                self.energy_unit_comboBox.setCurrentText(current_energy_unit)

            elif "T" in self.is_None:  # 知道T不知道G
                T = solve_for_T(kTST, Δn, σ, G)
                self.temp_lineEdit.setText(smart_format_float(T - 273.15, scientific_notation_limit=6))
                self.last_unknowns.append("T")

        # 不知道动力学，从kTST算时间、转化率
        if "G" not in self.is_None and "T" not in self.is_None:
            print(f"Calculating rate constant from TST.\n    Δn: {Δn}, σ: {σ}, T: {T} K, ΔG: {G} J/mol.")
            kTST = get_k_TST(Δn, σ, T, G)
            self.set_kTST_lineEdit(kTST)
            self.last_unknowns.append("kTST")
            if "t" in self.is_None:
                t = t_from_kinetics(kTST, conv, c1, c2)

                # 把时间单位换回秒，填入答案，再把单位切换到想要的单位，这样对应的slot会自动换算单位
                current_time_unit = self.time_unit_comboBox.currentText()
                self.time_unit_comboBox.setCurrentText("s")
                self.total_time_lineEdit.setText(smart_format_float(t))
                self.last_unknowns.append("t")
                self.time_unit_comboBox.setCurrentText(current_time_unit)

            elif 'conv' in self.is_None:
                conv = conv_from_kinetics(kTST, t, c1, c2)
                self.last_unknowns.append("conv")
                if conv == 1:
                    self.conversion_lineEdit.setText("~100")
                else:
                    self.conversion_lineEdit.setText(smart_format_float(conv * 100))

        self.last_unknowns = list(set(self.last_unknowns))
        assert len(self.last_unknowns)==2
        self.recalc_pushButton.setText("Recalc. "+self.RECALC_MAPPING[self.last_unknowns[0]] + " && " +\
                                       self.RECALC_MAPPING[self.last_unknowns[1]])
        print("---------------------------------\n\n")

    def automation(self,
                   input_data: Dict[str, str or Real],
                   units: Sequence[str],
                   missing_parameters: Sequence[str] = [],
                   redirect_sys_output = False):
        """

        Args:
            units: a 2-tuple [energy_unit (kJ/mol, kcal/mol, eV), time_unit (s, min, h, d, year)]
            input_data: a dict missing the correct parameters for calculation, the unit of the data is the unit used in the GUI

            missing_parameters: This parameter is only used in Verification.py,
                                Include to calculate and return which parameter, select from the standard key of standard keys

            redirect_sys_output: This is used in verification.py, to suppress excessive output when testing multiple cases

            (standard keys: mode, G, T, kTST, conc1, conc2, conv, t)

            e.g. missing kTST and t:    {
                                            "mode": "AB",
                                            "T": 25,
                                            "conc1": 0.5,
                                            "conc2": 1.2,
                                            "G": 15,
                                            "conv": 15
                                        },
        Returns:
            The value of the missed parameters

        """
        if redirect_sys_output:
            old_stdout = sys.stdout
            sys.stdout = None


        self.reset_all()
        self.conversion_lineEdit.setText("")

        self.energy_unit_comboBox.setCurrentText(units[0])
        self.time_unit_comboBox.setCurrentText(units[1])

        for key in input_data:
            if key == 'mode':
                self.MODE_MAPPING[input_data[key]].click()
            else:
                self.INPUT_MAPPING[key].setText(str(input_data[key]))

        # automation test case 有问题
        if not self.calculate_pushButton.isEnabled():
            print(input_data,units,missing_parameters)
            for key,value in self.MODE_MAPPING.items():
                print(key,repr(value.isChecked()))
            for key,value in self.INPUT_MAPPING.items():
                print(key,repr(value.text()))
            print("Energy Unit:",self.energy_unit_comboBox.currentText())
            print("Time Unit:",self.time_unit_comboBox.currentText())
            raise Exception("Calculate PushButton not enabled error.")

        self.calculate_pushButton.click()
        Application.processEvents()

        ret = []
        if missing_parameters:
            for i in missing_parameters:
                ret.append(self.INPUT_MAPPING[i].text())

        if redirect_sys_output:
            sys.stdout = old_stdout

        return ret

    def recalc(self):
        for i in self.last_unknowns:
            self.INPUT_MAPPING[i].setText("")
        if self.calculate_pushButton.isEnabled():
            self.calculate_pushButton.click()

    def batch(self):
        opened_xlsxs = get_open_file_UI(self,
                                       "",
                                       'xlsx',
                                       'Select batch input file created by filling the table from Batch_Template.xltx')
        for opened_xlsx in opened_xlsxs:
            self.run_batch(opened_xlsx)

    def run_batch(self, xlsx_input):
        """
        Run batch calculation from an xlsx input
        Args:
            xlsx_input: an xlsx input file created by filling Batch/Batch_Template.xltx

        Returns:

        """

        # build output xlsx
        xlsx_output = get_unused_filename(filename_class(os.path.realpath(xlsx_input)).replace_append_to("Erying_Eq_Output.xlsx"),
                                          use_proper_filename=False)
        xlsx_content = read_xlsx(xlsx_input)
        workbook = openpyxl.load_workbook(xlsx_input)
        worksheet = workbook[workbook.sheetnames[0]]


        HEADER_TO_DICT_KEY = {"Mode (A, AA, AB, Acat)": "mode",
                              "ΔG≠ (selected unit)": "G",
                              "ΔG≠ unit\n(kJ/mol, kcal/mol, eV)": "energy_unit",
                              "Temperature (°C)": "T",
                              "σ (default 1)": "σ",
                              "kTST (s-1 or M-1·s-1)": "kTST",
                              "Conc. 1 (mol/L)": "conc1",
                              "Conc. 2 (mol/L)": "conc2",
                              "Conversion (%)": "conv",
                              "Reaction time (selected unit)": "t",
                              "Reaction time unit\n(s, min, h, d, year)": "time_unit"}

        xlsx_content = transpose_2d_list(xlsx_content)  # 横纵交换，一行一个case
        HEADER_TO_DICT_KEY_XLSX_ORDER = OrderedDict()
        # reorder the header so it's in the same order as the input xlsx
        for header in xlsx_content[0]:
            HEADER_TO_DICT_KEY_XLSX_ORDER[header] = HEADER_TO_DICT_KEY[header]

        rets: List[List[str]] = [list(HEADER_TO_DICT_KEY_XLSX_ORDER.keys())]
        for case_count,test_case in enumerate(xlsx_content[1:]):
            xlsx_input_dict = {}
            for parameter_count, parameter in enumerate(xlsx_content[0]):
                xlsx_input_dict[HEADER_TO_DICT_KEY_XLSX_ORDER[parameter]] = test_case[parameter_count]
            units = [xlsx_input_dict.pop('energy_unit'), xlsx_input_dict.pop('time_unit')]

            self.automation(xlsx_input_dict, units)
            Application.processEvents()

            ret = [xlsx_input_dict.pop('mode')]
            for key in xlsx_input_dict.keys():
                ret.append(self.INPUT_MAPPING[key].text())
            HEADER_VALUE_LIST = list(HEADER_TO_DICT_KEY_XLSX_ORDER.values())
            ret.insert(HEADER_VALUE_LIST.index("energy_unit"), self.energy_unit_comboBox.currentText())
            ret.insert(HEADER_VALUE_LIST.index("time_unit"), self.time_unit_comboBox.currentText())

            for row_count,data in enumerate(ret):
                cell = worksheet.cell(row = row_count+1,column = case_count+2,value = data.strip(self.kTST_is_calculated_marker))
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')


        workbook.save(xlsx_output)
        open_explorer_and_select(xlsx_output)
        print(f"Batch processing of {xlsx_input} finished.")


if __name__ == '__main__':
    my_Qt_Program = myWidget()
    my_Qt_Program.show()

    sys.exit(Application.exec())
